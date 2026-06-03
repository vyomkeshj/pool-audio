#!/usr/bin/env python3
"""Build measurement_audio_index.json from the WAV folders + TEF protocol xlsx.

Scans every on-disk recording folder, parses each filename into its measurement
signature (M-flags, aeration, valve positions, noise, device, session), groups
files by signature, derives a pump-state label, and cross-references the protocol
spreadsheet rows (matched by saving_folder + parameters) for provenance.

Run:  python3 build_index.py
"""
import os
import re
import json
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(ROOT, "TEF_mic_test_protocol.xlsx")
OUT = os.path.join(ROOT, "measurement_audio_index.json")

# Folders that physically exist on disk (the recordings).
DISK_FOLDERS = [
    "testbed_motor_audio",
    "Testbedmotor",
    "Testbedmotor5_14",
    "Testbedmotor5_15",
    "testbedmotor5_25wav",
]

HEAD = re.compile(r"^(\d{8})_(\d{6})_(\d)_(\d)_(\d)_(\d)_(cam\d|mic\d)_(.*)", re.I)
TAIL = re.compile(
    r"_aeration_(\d+)_valveIn_(\d+)_valveOut_(\d+)_noise_([A-Za-z0-9]+)_(\d+)\.wav$",
    re.I,
)

NOISE_LEGEND = {
    "N": "clean / no added background noise (reference)",
    "A": "children's playground (detske hriste)",
    "B": "lawnmower (sekacka) - petrol/electric",
    "C": "road traffic (doprava)",
    "D": "human speech (lidska rec)",
    "E": "music / song (pisnicka)",
}


def norm_folder(name):
    """Normalise a folder name so spreadsheet and disk spellings match."""
    n = str(name).strip().lower()
    if n.endswith("wav"):
        n = n[:-3]
    return n.rstrip("_")


def parse_filename(fn):
    h = HEAD.match(fn)
    t = TAIL.search(fn)
    if not h or not t:
        return None
    return {
        "date": h.group(1),
        "time": h.group(2),
        "timestamp": f"{h.group(1)}_{h.group(2)}",
        "M1": int(h.group(3)),
        "M2": int(h.group(4)),
        "M3": int(h.group(5)),
        "M4": int(h.group(6)),
        "device": h.group(7).lower(),
        "aeration": int(t.group(1)),
        "valveIn": int(t.group(2)),
        "valveOut": int(t.group(3)),
        "noise": t.group(4).upper(),
        "repetition": int(t.group(5)),
    }


def signature(p):
    return (
        f"M2{p['M2']}_M3{p['M3']}_M4{p['M4']}"
        f"_aer{p['aeration']}_vin{p['valveIn']}_vout{p['valveOut']}_noise{p['noise']}"
    )


def derive_state(p):
    """Map raw parameters to pump-state labels for detection work."""
    aux = [m for m in ("M2", "M3", "M4") if p[m] == 1]
    labels = []
    if p["aeration"] >= 1:
        labels.append("aerating")
    if p["valveIn"] > 1:
        labels.append(f"suction_blockage_L{p['valveIn']}")
    if p["valveOut"] > 1:
        labels.append(f"discharge_blockage_L{p['valveOut']}")
    if aux:
        labels.append("multi_pump_" + "_".join(aux))
    if not labels:
        labels.append("normal")

    # Single best summary label (priority: aeration > blockage > multi-pump > normal).
    if p["aeration"] >= 1:
        nominal = "aerating"
    elif p["valveIn"] > 1:
        nominal = "suction_blockage"
    elif p["valveOut"] > 1:
        nominal = "discharge_blockage"
    elif aux:
        nominal = "multi_pump"
    else:
        nominal = "normal"

    return {
        "nominal_state": nominal,
        "labels": labels,
        "aux_pumps_on": aux,
        "blockage_in_level": p["valveIn"],
        "blockage_out_level": p["valveOut"],
        "aeration_on": bool(p["aeration"]),
    }


def load_protocol():
    """Return {norm_folder: {param_key: [excel_rows]}} from List1."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["List1"]
    out = {}
    rows_by_folder = {}
    for r in range(4, ws.max_row + 1):
        m2, m3, m4 = ws.cell(r, 4).value, ws.cell(r, 5).value, ws.cell(r, 6).value
        aer = ws.cell(r, 7).value
        vin, vout = ws.cell(r, 8).value, ws.cell(r, 9).value
        noise = ws.cell(r, 10).value
        folder = ws.cell(r, 14).value
        if folder is None or vin is None or vout is None or noise is None:
            continue

        def to_int(x):
            try:
                return int(x)
            except (TypeError, ValueError):
                return 0

        nf = norm_folder(folder)
        key = (
            to_int(m2), to_int(m3), to_int(m4), to_int(aer),
            to_int(vin), to_int(vout), str(noise).strip().upper(),
        )
        out.setdefault(nf, {}).setdefault(key, []).append(r)
        rows_by_folder.setdefault(nf, {"raw_name": str(folder), "rows": []})["rows"].append(r)
    return out, rows_by_folder


def main():
    protocol, prot_folders = load_protocol()

    measurements = {}
    folder_summaries = {}
    total_files = 0

    for folder in DISK_FOLDERS:
        nf = norm_folder(folder)
        fdir = os.path.join(ROOT, folder)
        files = sorted(os.listdir(fdir))
        folder_summaries[folder] = {"n_files": len(files), "n_signatures": 0}

        # group: signature -> session(timestamp,rep) -> {device: filename}
        for fn in files:
            p = parse_filename(fn)
            if p is None:
                raise ValueError(f"unparseable filename: {fn}")
            total_files += 1
            sig = signature(p)

            entry = measurements.setdefault(sig, {
                "params": {k: p[k] for k in
                           ("M1", "M2", "M3", "M4", "aeration", "valveIn", "valveOut", "noise")},
                "state": derive_state(p),
                "protocol_excel_rows": {},
                "n_files": 0,
                "folders": [],
                "_sessions": {},  # (folder,timestamp) -> session dict
            })
            entry["n_files"] += 1
            if folder not in entry["folders"]:
                entry["folders"].append(folder)

            skey = (folder, p["timestamp"])
            sess = entry["_sessions"].setdefault(skey, {
                "folder": folder,
                "timestamp": p["timestamp"],
                "repetition": p["repetition"],
                "n_devices": 0,
                "devices": {},
            })
            sess["devices"][p["device"]] = fn
            sess["n_devices"] = len(sess["devices"])

            # attach protocol excel rows for this folder+params
            pkey = (p["M2"], p["M3"], p["M4"], p["aeration"], p["valveIn"], p["valveOut"], p["noise"])
            rows = protocol.get(nf, {}).get(pkey)
            if rows:
                entry["protocol_excel_rows"][folder] = rows

    # finalise: turn session dict into a sorted list, count signatures per folder
    sig_per_folder = {f: set() for f in DISK_FOLDERS}
    for sig, entry in measurements.items():
        sessions = sorted(entry.pop("_sessions").values(),
                          key=lambda s: (s["folder"], s["timestamp"]))
        entry["sessions"] = sessions
        entry["n_sessions"] = len(sessions)
        for f in entry["folders"]:
            sig_per_folder[f].add(sig)
    for f in DISK_FOLDERS:
        folder_summaries[f]["n_signatures"] = len(sig_per_folder[f])

    # folders referenced by the protocol but absent on disk
    disk_norm = {norm_folder(f) for f in DISK_FOLDERS}
    missing = {prot_folders[nf]["raw_name"]: len(prot_folders[nf]["rows"])
               for nf in prot_folders if nf not in disk_norm}

    out = {
        "_meta": {
            "source_spreadsheet": "TEF_mic_test_protocol.xlsx",
            "sheet": "List1",
            "audio_dirs": DISK_FOLDERS,
            "description": (
                "Maps each recorded measurement signature to its matching .wav files "
                "across all on-disk folders. ALL equipment is healthy -- nothing is "
                "broken; signatures are operating configurations, not faults. "
                "M1+M2 are large pumps (M1 always on), M3+M4 are exhaust fans, "
                "aeration is an air injector; valveIn throttles M1's suction side and "
                "valveOut M1's discharge side (1 = open, higher = more restricted). "
                "Signature encodes M2/M3/M4 on/off, aeration, valveIn, valveOut, noise. "
                "Each session = one recording instance (timestamp + repetition) "
                "capturing 16 devices (cam1-8 + mic1-8). Signatures are merged across "
                "folders; each session carries its source folder. NOTE: the legacy "
                "'nominal_state' field collapses two orthogonal axes (M1 flow "
                "restriction vs which auxiliary machine is running) into one label and "
                "is kept only for back-compat -- prefer the raw params + labels[]."
            ),
            "audio_format": "mono 16-bit PCM WAV, 44.1 kHz, ~60 s per file",
            "n_signatures": len(measurements),
            "n_files_total": total_files,
            "filename_fields":
                "YYYYMMDD_HHMMSS_M1_M2_M3_M4_<device>_<deviceDesc>_M1_aeration_A_"
                "valveIn_I_valveOut_O_noise_T_REP.wav",
            "protocol_folders_missing_on_disk": missing,
            "noise_legend": NOISE_LEGEND,
            "equipment_legend": {
                "M1": "large pump, always ON (the pump being monitored)",
                "M2": "second large pump, on/off (healthy)",
                "M3": "exhaust fan, on/off (healthy; adds ~4-8 kHz broadband airflow)",
                "M4": "exhaust fan, on/off (healthy; near-silent)",
                "aeration": "air injector, on/off (only in 5_25; on-clips are anomalous)",
                "valveIn": "M1 suction-side throttle, 1=open .. 4=most restricted",
                "valveOut": "M1 discharge-side throttle, level in {1,2,3,4,5,8,11}, 1=open",
            },
            "state_legend_DEPRECATED": {
                "_note": "Legacy single-label collapse of two orthogonal axes; nothing "
                         "is broken. Kept for back-compat -- prefer raw params + labels[].",
                "normal": "M1 only, valves open (in=1,out=1), no aeration - baseline",
                "suction_blockage": "valveIn>1 - M1 inlet throttled (NOT a pump fault)",
                "discharge_blockage": "valveOut>1 - M1 outlet throttled; level = severity",
                "aerating": "aeration on (air injector running)",
                "multi_pump": "an auxiliary machine (M2 pump / M3,M4 fans) also running",
            },
        },
        "folder_summary": folder_summaries,
        "measurements": dict(sorted(measurements.items())),
    }

    with open(OUT, "w") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    print(f"wrote {OUT}")
    print(f"  signatures: {len(measurements)}  files: {total_files}")
    print(f"  folders: {[(f, folder_summaries[f]['n_files']) for f in DISK_FOLDERS]}")
    print(f"  protocol folders missing on disk: {missing}")
    # state distribution
    from collections import Counter
    st = Counter()
    for e in measurements.values():
        st[e["state"]["nominal_state"]] += e["n_files"]
    print(f"  files per nominal state: {dict(st)}")


if __name__ == "__main__":
    main()
