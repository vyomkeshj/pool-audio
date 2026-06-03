#!/usr/bin/env python3
"""Build a fresh measurement index over ALL on-disk recording folders, including
the two newly-added campaigns (testbedmotor5_19wav, testbedmotor5_20wav) that
extend the suction (valveIn) sweep to levels {5, 8, 11}.

Self-contained (does not depend on the parent build_index.py). Scans every
folder, parses each filename into its measurement parameters, groups files into
sessions (one timestamp = 16 simultaneous device channels), and cross-references
the protocol spreadsheet rows for provenance.

Output: run/measurement_index.json

Run: python3 run/build_index.py
"""
import os
import re
import json
from collections import Counter, defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)                      # pool-audio/
XLSX = os.path.join(ROOT, "TEF_mic_test_protocol.xlsx")
OUT = os.path.join(HERE, "measurement_index.json")

# (disk_path_relative_to_ROOT, canonical_folder_label, protocol_norm_name)
# The 5_19 / 5_20 archives nest the wavs one level deep.
FOLDERS = [
    ("testbed_motor_audio",                    "testbed_motor_audio", "testbed_motor_audio"),
    ("Testbedmotor",                           "Testbedmotor",        "testbedmotor"),
    ("Testbedmotor5_14",                       "Testbedmotor5_14",    "testbedmotor5_14"),
    ("Testbedmotor5_15",                       "Testbedmotor5_15",    "testbedmotor5_15"),
    ("testbedmotor5_19wav/testbedmotor5_19wav","testbedmotor5_19",    "testbedmotor5_19"),
    ("testbedmotor5_20wav/testbedmotor5_20wav","testbedmotor5_20",    "testbedmotor5_20"),
    ("testbedmotor5_25wav",                    "testbedmotor5_25",    "testbedmotor5_25"),
]

HEAD = re.compile(r"^(\d{8})_(\d{6})_(\d)_(\d)_(\d)_(\d)_(cam\d|mic\d)_(.*)", re.I)
TAIL = re.compile(
    r"_aeration_(\d+)_valveIn_(\d+)_valveOut_(\d+)_n?oise_([A-Za-z0-9]+)_(\d+)\.wav$",
    re.I,
)

NOISE_LEGEND = {
    "N": "clean / no added background noise (reference)",
    "A": "children's playground", "B": "lawnmower", "C": "road traffic",
    "D": "human speech", "E": "music / song",
}


def parse_filename(fn):
    h = HEAD.match(fn)
    t = TAIL.search(fn)
    if not h or not t:
        return None
    return {
        "timestamp": f"{h.group(1)}_{h.group(2)}",
        "M1": int(h.group(3)), "M2": int(h.group(4)),
        "M3": int(h.group(5)), "M4": int(h.group(6)),
        "device": h.group(7).lower(),
        "aeration": int(t.group(1)), "valveIn": int(t.group(2)),
        "valveOut": int(t.group(3)), "noise": t.group(4).upper(),
        "repetition": int(t.group(5)),
    }


def signature(p):
    return (f"M2{p['M2']}_M3{p['M3']}_M4{p['M4']}_aer{p['aeration']}"
            f"_vin{p['valveIn']}_vout{p['valveOut']}_noise{p['noise']}")


def load_protocol():
    """Return {protocol_norm_name: {param_key: [rows]}} from sheet List1."""
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["List1"]
    out = {}

    def to_int(x):
        try:
            return int(x)
        except (TypeError, ValueError):
            return 0

    def norm(name):
        n = str(name).strip().lower()
        if n.endswith("wav"):
            n = n[:-3]
        return n.rstrip("_")

    for r in range(4, ws.max_row + 1):
        m2, m3, m4 = ws.cell(r, 4).value, ws.cell(r, 5).value, ws.cell(r, 6).value
        aer = ws.cell(r, 7).value
        vin, vout = ws.cell(r, 8).value, ws.cell(r, 9).value
        noise = ws.cell(r, 10).value
        folder = ws.cell(r, 14).value
        if folder is None or vin is None or vout is None or noise is None:
            continue
        key = (to_int(m2), to_int(m3), to_int(m4), to_int(aer),
               to_int(vin), to_int(vout), str(noise).strip().upper())
        out.setdefault(norm(folder), {}).setdefault(key, []).append(r)
    return out


def main():
    protocol = load_protocol()
    measurements = {}
    folder_summaries = {}
    total = 0
    bad = []

    for disk, label, pnorm in FOLDERS:
        fdir = os.path.join(ROOT, disk)
        files = sorted(f for f in os.listdir(fdir) if f.lower().endswith(".wav"))
        folder_summaries[label] = {"n_files": len(files), "disk_path": disk}
        for fn in files:
            p = parse_filename(fn)
            if p is None:
                bad.append(fn)
                continue
            total += 1
            sig = signature(p)
            entry = measurements.setdefault(sig, {
                "params": {k: p[k] for k in
                           ("M1", "M2", "M3", "M4", "aeration",
                            "valveIn", "valveOut", "noise")},
                "protocol_excel_rows": {},
                "n_files": 0, "folders": [], "_sessions": {},
            })
            entry["n_files"] += 1
            if label not in entry["folders"]:
                entry["folders"].append(label)
            skey = (label, p["timestamp"])
            sess = entry["_sessions"].setdefault(skey, {
                "folder": label, "disk_path": disk,
                "timestamp": p["timestamp"], "repetition": p["repetition"],
                "devices": {},
            })
            sess["devices"][p["device"]] = fn

            pkey = (p["M2"], p["M3"], p["M4"], p["aeration"],
                    p["valveIn"], p["valveOut"], p["noise"])
            rows = protocol.get(pnorm, {}).get(pkey)
            if rows:
                entry["protocol_excel_rows"][label] = rows

    for sig, e in measurements.items():
        sessions = sorted(e.pop("_sessions").values(),
                          key=lambda s: (s["folder"], s["timestamp"]))
        for s in sessions:
            s["n_devices"] = len(s["devices"])
        e["sessions"] = sessions
        e["n_sessions"] = len(sessions)

    out = {
        "_meta": {
            "source_spreadsheet": "TEF_mic_test_protocol.xlsx",
            "audio_format": "mono 16-bit PCM WAV, 44.1 kHz, ~60 s",
            "folders": [f[1] for f in FOLDERS],
            "n_signatures": len(measurements),
            "n_files_total": total,
            "n_unparseable": len(bad),
            "noise_legend": NOISE_LEGEND,
            "note": ("ALL equipment healthy; signatures are operating configs, "
                     "not faults. valveIn=suction throttle, valveOut=discharge "
                     "throttle (1=open, higher=more restricted). New 5_19/5_20 "
                     "campaigns extend valveIn to {5,8,11}."),
        },
        "folder_summary": folder_summaries,
        "measurements": dict(sorted(measurements.items())),
    }
    with open(OUT, "w") as fh:
        json.dump(out, fh, indent=2, ensure_ascii=False)

    print(f"wrote {OUT}")
    print(f"  signatures={len(measurements)}  files={total}  unparseable={len(bad)}")
    for label, s in folder_summaries.items():
        print(f"    {label:24s} {s['n_files']:5d}")
    # valve coverage among M1-only (M2=M3=M4=aer=0) sessions
    vin = Counter(); vout = Counter()
    m1_sessions = 0
    for e in measurements.values():
        pr = e["params"]
        if pr["M2"] or pr["M3"] or pr["M4"] or pr["aeration"]:
            continue
        m1_sessions += e["n_sessions"]
        vin[pr["valveIn"]] += e["n_sessions"]
        vout[pr["valveOut"]] += e["n_sessions"]
    print(f"  M1-only sessions={m1_sessions}")
    print(f"  valveIn  sessions: {dict(sorted(vin.items()))}")
    print(f"  valveOut sessions: {dict(sorted(vout.items()))}")
    if bad:
        print(f"  unparseable examples: {bad[:3]}")


if __name__ == "__main__":
    main()
